from flask import Flask, render_template, request, send_file, jsonify
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import json
import io
import tempfile

app = Flask(__name__, template_folder='../templates', static_folder='../static')

IDR_COLUMNS = [
    'Item Group', 'Item Name', 'Item Parent', 'Conversion Factor',
    'Projection Units', 'Total KGs', 'Packaging Type', 'Packaging Method',
    'Item Type', 'Customer Group', 'Origin', 'Created By',
    'Month', 'Year', 'Last Modified Date', 'Last Modified Time',
    'BOM Item Type', 'Customer', 'Key Account Manager', 'Item Code'
]

DEPT_DATA = [
    ('Production', 80, 48),
    ('Airlines', 34, 14),
    ('DB Conversion', 48, ''),
    ('Conti', 12, ''),
    ('PFS', 16, 16),
    ('FW', 36, 12),
    ('DB Bulk', 35, ''),
    ('VA Bulk', 17, 12),
    ('Inventory', 15, 5),
    ('Dispatch', 20, 8),
    ('RTV', 10, ''),
    ('Quality', 12, 3),
    ('FFS', 10, 10),
    ('HK', 16, 2),
    ('R&M', 5, ''),
]

SHIFT_CAPACITY = {
    'Conti Mix': 6500,
    'FW': '=90*60*3*7',
    'Jar Sealing': 2500,
    'Jar Sealing - WAD Machine': 2400,
    'LD Pack of 2': 2500,
    'LD Pack of 4': '=1248/0.8',
    'Pace - FFS': '=80*60*7',
    'Perfect - FFS': '=45*60*6',
    'PFS': 24000,
    'Table': 2500,
    'Table - Airlines': 18000,
}

ROASTING_LOOKUP = {
    'Apple Pie Date Bites': (0.401, 460),
    'Berry Mix': (0, 0),
    'Classic Date Bites': (0.4555, 262),
    'Classic Date bites': (0.4555, 262),
    'Coffe Rush Date Bites': (0.37, 460),
    'Dark Choco Date Bites': (0.401, 460),
    'Fruit and Nut Mix': (0.20, 300),
    'Mexican Peri Peri Snack Mix': (0.54, 600),
    'Mexican Peri Peri Snack Mix (FFS)': (0.70, 600),
    'Nut Mix': (1.0, 250),
    'Nut Mix - Airlines': (1.0, 250),
    'Premium Panchmewa Superfood': (0.40, 480),
    'Premium Panchmewa Superfood (FFS)': (0.30, 600),
    'Roasted and Salted Almonds': (1.0, 250),
    'Roasted and Salted Cashew': (1.0, 250),
    'Roasted and Salted Cashew (W240)': (1.0, 250),
    'Satva Mix': (0.63, 300),
    'Seed Mix': (0.875, 600),
    'Smokey Almonds': (0.92, 250),
    'Smokey Almonds - Airlines': (1.0, 250),
    'Smokey BBQ Party Mix': (0.94, 600),
    'Sweet and Salty Mix': (0.20, 300),
    'Trail Mix': (0.59, 600),
    'Tropical Savoury Mix': (0.23, 600),
    'Cream N Onion Makhana': (1.0, 100),
    'Mango Flavoured Raisins': (1.0, 200),
    'Roasted Black Pepper Cashew': (1.0, 240),
    'Roasted Chatpata Cashew': (1.0, 240),
    'Roasted Thai Chilli Cashew': (1.0, 240),
    'Roasted & Salted Pistachios': (1.0, 250),
    'Jumbo Iranian Roasted & Salted Pistachios': (1.0, 250),
    'Prasadam Roasted Makhana': (1.0, 0),
    'Paan Mix': (0, 0),
}

FFS_METHODS = {'Pace - FFS', 'Perfect - FFS'}

PACKAGING_OVERRIDES = {
    'Dark Choco-Orange Date Bite Farmley Tin Jar 200g': {'Packaging Type': 'Pouch', 'Packaging Method': 'FW'},
}


def process_files(this_month_file, lookup_file, cf_value, working_days):
    xls = pd.ExcelFile(this_month_file, engine='openpyxl')
    proj_sheet = None
    for name in ['Projection', 'Query Report']:
        if name in xls.sheet_names:
            proj_sheet = name
            break
    if proj_sheet is None:
        proj_sheet = xls.sheet_names[0]
    df_proj = pd.read_excel(xls, sheet_name=proj_sheet)
    df_proj.columns = [c.strip() for c in df_proj.columns]

    lookup_file.seek(0)
    df_lookup = pd.read_excel(lookup_file, engine='openpyxl')
    df_lookup.columns = [c.strip() for c in df_lookup.columns]

    lookup = {}
    for _, row in df_lookup.iterrows():
        name = str(row.get('Item Name', '')).strip()
        if name and name != 'nan':
            if name not in lookup:
                lookup[name] = {
                    'Packaging Type': str(row.get('Packaging Type', '')).strip(),
                    'Packaging Method': str(row.get('Packaging Method', '')).strip()
                }

    lookup.update(PACKAGING_OVERRIDES)

    idr_rows = []
    for _, row in df_proj.iterrows():
        item_name = str(row.get('Item Name', '')).strip()
        pkg = lookup.get(item_name, {'Packaging Type': '', 'Packaging Method': ''})
        idr_rows.append({
            'Item Group': row.get('Item Group', ''),
            'Item Name': row.get('Item Name', ''),
            'Item Parent': row.get('Item Parent', ''),
            'Conversion Factor': row.get('Conversion Factor', ''),
            'Projection Units': row.get('Projection Units', ''),
            'Total KGs': row.get('Total KGs', ''),
            'Packaging Type': pkg['Packaging Type'],
            'Packaging Method': pkg['Packaging Method'],
            'Item Type': row.get('Item Type', ''),
            'Customer Group': row.get('Customer Group', ''),
            'Origin': row.get('Origin', row.get('Warehouse', '')),
            'Created By': row.get('Created By', ''),
            'Month': row.get('Month', ''),
            'Year': row.get('Year', ''),
            'Last Modified Date': row.get('Last Modified Date', ''),
            'Last Modified Time': row.get('Last Modified Time', ''),
            'BOM Item Type': row.get('BOM Item Type', ''),
            'Customer': row.get('Customer', ''),
            'Key Account Manager': row.get('Key Account Manager', ''),
            'Item Code': row.get('Item Code', '')
        })

    df_idr = pd.DataFrame(idr_rows, columns=IDR_COLUMNS)
    df_idr = df_idr[df_idr['Origin'] == 'Indore'].reset_index(drop=True)
    df_idr['Projection Units'] = pd.to_numeric(df_idr['Projection Units'], errors='coerce').fillna(0)
    df_idr['Total KGs'] = pd.to_numeric(df_idr['Total KGs'], errors='coerce').fillna(0)

    valid_pkg = df_idr['Packaging Method'].notna() & (df_idr['Packaging Method'] != '') & (df_idr['Packaging Method'] != 'nan')
    pivot_source = df_idr[valid_pkg].groupby('Packaging Method').agg(
        **{'Projected Units': ('Projection Units', 'sum'),
           'Projected Kg': ('Total KGs', 'sum')}
    ).reset_index().rename(columns={'Packaging Method': 'Source'})

    fw_mask = df_idr['Packaging Method'] == 'FW'
    pivot_fw = df_idr[fw_mask].groupby('Item Name').agg(
        **{'Projected Units': ('Projection Units', 'sum'),
           'Projected Kg': ('Total KGs', 'sum')}
    ).reset_index()

    ffs_mask = df_idr['Packaging Method'].isin(FFS_METHODS)
    pivot_ffs = df_idr[ffs_mask].groupby('Item Name').agg(
        **{'Projected Units': ('Projection Units', 'sum'),
           'Projected Kg': ('Total KGs', 'sum')}
    ).reset_index()

    va_mask = df_idr['Item Type'] == 'Value Added'
    pivot_va = df_idr[va_mask].groupby('Item Parent').agg(
        **{'Projected Kg': ('Total KGs', 'sum')}
    ).reset_index().rename(columns={'Item Parent': 'Item Name'})

    return df_proj, df_idr, pivot_source, pivot_fw, pivot_ffs, pivot_va


def build_output_excel(df_proj, df_idr, pivot_source, pivot_fw, pivot_ffs, pivot_va, cf_value, working_days):
    wb = openpyxl.Workbook()
    hf = Font(bold=True, color='FFFFFF', size=11)
    hfill = PatternFill('solid', fgColor='2F5496')
    ha = Alignment(horizontal='center', vertical='center', wrap_text=True)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    nk = '#,##0.00'
    nu = '#,##0'
    bf = Font(bold=True)
    sfill = PatternFill('solid', fgColor='D6E4F0')
    yfill = PatternFill('solid', fgColor='FFFF00')
    lbl_font = Font(bold=True, size=11, color='2F5496')

    # --- Tab 1: Projection (raw input) ---
    ws_raw = wb.active
    ws_raw.title = 'Projection'
    proj_cols = list(df_proj.columns)
    for ci, cn in enumerate(proj_cols, 1):
        c = ws_raw.cell(row=1, column=ci, value=cn)
        c.font = hf; c.fill = hfill; c.alignment = ha; c.border = tb
    for ri, (_, row) in enumerate(df_proj.iterrows(), 2):
        for ci, cn in enumerate(proj_cols, 1):
            v = row[cn]
            if pd.isna(v): v = ''
            ws_raw.cell(row=ri, column=ci, value=v).border = tb
    for ci in range(1, len(proj_cols)+1):
        ws_raw.column_dimensions[get_column_letter(ci)].width = 18
    ws_raw.auto_filter.ref = f"A1:{get_column_letter(len(proj_cols))}{len(df_proj)+1}"
    ws_raw.freeze_panes = 'A2'

    # --- Tab 2: IDR Projection ---
    ws_idr = wb.create_sheet('IDR Projection')
    for ci, cn in enumerate(IDR_COLUMNS, 1):
        c = ws_idr.cell(row=1, column=ci, value=cn)
        c.font = hf; c.fill = hfill; c.alignment = ha; c.border = tb
    for ri, (_, row) in enumerate(df_idr.iterrows(), 2):
        for ci, cn in enumerate(IDR_COLUMNS, 1):
            v = row[cn]
            if pd.isna(v): v = ''
            c = ws_idr.cell(row=ri, column=ci, value=v)
            c.border = tb
            if cn in ('Total KGs', 'Conversion Factor'): c.number_format = nk
            elif cn == 'Projection Units': c.number_format = nu
    for ci in range(1, len(IDR_COLUMNS)+1):
        ws_idr.column_dimensions[get_column_letter(ci)].width = 18
    ws_idr.auto_filter.ref = f"A1:T{len(df_idr)+1}"
    ws_idr.freeze_panes = 'A2'

    # --- Tab 3: Pivot ---
    ws_p = wb.create_sheet('Pivot')

    # ========== SECTION 1: Source Summary (A-H) ==========
    ws_p.cell(row=2, column=4, value=cf_value).font = bf
    ws_p.cell(row=2, column=6, value=working_days).font = bf

    s1_headers = ['Source', 'Projected Units', 'Projected Kg', '1.3 CF',
                  'Sum of Pro (1.3 CF) Kg', 'Per Day PCS Prod.',
                  'No. of PCS Prod. (per shift)', 'No of Cycles']
    for ci, h in enumerate(s1_headers, 1):
        c = ws_p.cell(row=3, column=ci, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ha; c.border = tb

    s1_start = 4
    for ri, (_, row) in enumerate(pivot_source.iterrows()):
        r = s1_start + ri
        ws_p.cell(row=r, column=1, value=row['Source']).border = tb
        ws_p.cell(row=r, column=2, value=row['Projected Units']).border = tb
        ws_p.cell(row=r, column=2).number_format = nu
        kg = row['Projected Kg']
        if pd.notna(kg) and kg != 0:
            ws_p.cell(row=r, column=3, value=kg).border = tb
            ws_p.cell(row=r, column=3).number_format = nk
        else:
            ws_p.cell(row=r, column=3).border = tb
        ws_p.cell(row=r, column=4, value=f'=B{r}*$D$2').border = tb
        ws_p.cell(row=r, column=4).number_format = nk
        ws_p.cell(row=r, column=5, value=f'=C{r}*$D$2').border = tb
        ws_p.cell(row=r, column=5).number_format = nk
        ws_p.cell(row=r, column=6, value=f'=D{r}/$F$2').border = tb
        ws_p.cell(row=r, column=6).number_format = nk
        cap = SHIFT_CAPACITY.get(row['Source'], '')
        ws_p.cell(row=r, column=7, value=cap).border = tb
        ws_p.cell(row=r, column=8, value=f'=F{r}/G{r}').border = tb
        ws_p.cell(row=r, column=8).number_format = nk

    s1_end = s1_start + len(pivot_source) - 1
    gt1 = s1_end + 1
    ws_p.cell(row=gt1, column=1, value='Grand Total').font = bf
    ws_p.cell(row=gt1, column=1).border = tb
    for ci in range(2, 9):
        cl = get_column_letter(ci)
        if ci == 7:
            ws_p.cell(row=gt1, column=ci, value=f'=E{gt1}').border = tb
        else:
            ws_p.cell(row=gt1, column=ci, value=f'=SUM({cl}{s1_start}:{cl}{s1_end})').border = tb
        ws_p.cell(row=gt1, column=ci).font = bf
        ws_p.cell(row=gt1, column=ci).number_format = nk

    # ========== SECTION 2a: FW Packaging Method (K-O) ==========
    ws_p.cell(row=1, column=11, value='Packaging Method').font = lbl_font
    ws_p.cell(row=1, column=12, value='FW').font = lbl_font
    ws_p.cell(row=1, column=12).fill = sfill

    ws_p.cell(row=2, column=14, value=cf_value).font = bf
    ws_p.cell(row=2, column=15, value=working_days).font = bf

    fw_headers = ['Item Name', 'Projected Units', 'Projected Kg', '1.3 CF', 'Per Day Production Kg']
    for ci, h in enumerate(fw_headers):
        c = ws_p.cell(row=3, column=11+ci, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ha; c.border = tb

    fw_start = 4
    for ri, (_, row) in enumerate(pivot_fw.iterrows()):
        r = fw_start + ri
        ws_p.cell(row=r, column=11, value=row['Item Name']).border = tb
        ws_p.cell(row=r, column=12, value=row['Projected Units']).border = tb
        ws_p.cell(row=r, column=12).number_format = nu
        ws_p.cell(row=r, column=13, value=row['Projected Kg']).border = tb
        ws_p.cell(row=r, column=13).number_format = nk
        ws_p.cell(row=r, column=14, value=f'=M{r}*$N$2').border = tb
        ws_p.cell(row=r, column=14).number_format = nk
        ws_p.cell(row=r, column=15, value=f'=N{r}/$O$2').border = tb
        ws_p.cell(row=r, column=15).number_format = nk

    fw_end = fw_start + max(len(pivot_fw) - 1, 0)
    fw_gt = fw_end + 1
    ws_p.cell(row=fw_gt, column=11, value='Grand Total').font = bf
    ws_p.cell(row=fw_gt, column=11).border = tb
    for ci in range(12, 16):
        cl = get_column_letter(ci)
        if len(pivot_fw) > 0:
            ws_p.cell(row=fw_gt, column=ci, value=f'=SUM({cl}{fw_start}:{cl}{fw_end})').border = tb
        else:
            ws_p.cell(row=fw_gt, column=ci, value=0).border = tb
        ws_p.cell(row=fw_gt, column=ci).font = bf
        ws_p.cell(row=fw_gt, column=ci).number_format = nk

    # ========== SECTION 2b: (Multiple Items) = Pace-FFS + Perfect-FFS (K-O) ==========
    ffs_label_row = fw_gt + 4
    ws_p.cell(row=ffs_label_row, column=11, value='Packaging Method').font = lbl_font
    ws_p.cell(row=ffs_label_row, column=12, value='(Multiple Items)').font = lbl_font
    ws_p.cell(row=ffs_label_row, column=12).fill = sfill

    ffs_param_row = ffs_label_row + 1
    ws_p.cell(row=ffs_param_row, column=14, value=cf_value).font = bf
    ws_p.cell(row=ffs_param_row, column=15, value=working_days).font = bf

    ffs_hdr_row = ffs_param_row + 1
    for ci, h in enumerate(fw_headers):
        c = ws_p.cell(row=ffs_hdr_row, column=11+ci, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ha; c.border = tb

    ffs_start = ffs_hdr_row + 1
    n2_ref = f'$N${ffs_param_row}'
    o2_ref = f'$O${ffs_param_row}'
    for ri, (_, row) in enumerate(pivot_ffs.iterrows()):
        r = ffs_start + ri
        ws_p.cell(row=r, column=11, value=row['Item Name']).border = tb
        ws_p.cell(row=r, column=12, value=row['Projected Units']).border = tb
        ws_p.cell(row=r, column=12).number_format = nu
        ws_p.cell(row=r, column=13, value=row['Projected Kg']).border = tb
        ws_p.cell(row=r, column=13).number_format = nk
        ws_p.cell(row=r, column=14, value=f'=M{r}*{n2_ref}').border = tb
        ws_p.cell(row=r, column=14).number_format = nk
        ws_p.cell(row=r, column=15, value=f'=N{r}/{o2_ref}').border = tb
        ws_p.cell(row=r, column=15).number_format = nk

    ffs_end = ffs_start + max(len(pivot_ffs) - 1, 0)
    ffs_gt = ffs_end + 1
    ws_p.cell(row=ffs_gt, column=11, value='Grand Total').font = bf
    ws_p.cell(row=ffs_gt, column=11).border = tb
    for ci in range(12, 16):
        cl = get_column_letter(ci)
        if len(pivot_ffs) > 0:
            ws_p.cell(row=ffs_gt, column=ci, value=f'=SUM({cl}{ffs_start}:{cl}{ffs_end})').border = tb
        else:
            ws_p.cell(row=ffs_gt, column=ci, value=0).border = tb
        ws_p.cell(row=ffs_gt, column=ci).font = bf
        ws_p.cell(row=ffs_gt, column=ci).number_format = nk

    # ========== SECTION 3: Value Added (R-Y) ==========
    ws_p.cell(row=1, column=18, value='Item Type').font = lbl_font
    ws_p.cell(row=1, column=19, value='Value Added').font = lbl_font
    ws_p.cell(row=1, column=19).fill = sfill

    ws_p.cell(row=2, column=20, value=cf_value).font = bf
    ws_p.cell(row=2, column=21, value=working_days).font = bf

    va_headers = ['Item Name', 'Projected Kg', '1.3 CF', 'Per Day Kg Prod',
                  'Roasting Item (%)', 'Roasting Kg', 'Oven Cycle (SKU Level)', 'No of Oven Cycle']
    for ci, h in enumerate(va_headers):
        c = ws_p.cell(row=3, column=18+ci, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ha; c.border = tb

    va_start = 4
    for ri, (_, row) in enumerate(pivot_va.iterrows()):
        r = va_start + ri
        item = row['Item Name']
        ws_p.cell(row=r, column=18, value=item).border = tb
        ws_p.cell(row=r, column=19, value=row['Projected Kg']).border = tb
        ws_p.cell(row=r, column=19).number_format = nk
        ws_p.cell(row=r, column=20, value=f'=S{r}*$T$2').border = tb
        ws_p.cell(row=r, column=20).number_format = nk
        ws_p.cell(row=r, column=21, value=f'=T{r}/$U$2').border = tb
        ws_p.cell(row=r, column=21).number_format = nk

        ws_p.cell(row=r, column=22).border = tb
        ws_p.cell(row=r, column=22).number_format = '0%'
        ws_p.cell(row=r, column=23, value=f'=U{r}*V{r}').border = tb
        ws_p.cell(row=r, column=23).number_format = nk
        ws_p.cell(row=r, column=24).border = tb
        ws_p.cell(row=r, column=24).number_format = nu
        ws_p.cell(row=r, column=25, value=f'=IF(X{r}>0,W{r}/X{r},0)').border = tb
        ws_p.cell(row=r, column=25).number_format = '0.00'

    va_end = va_start + max(len(pivot_va) - 1, 0)
    va_gt = va_end + 1
    ws_p.cell(row=va_gt, column=18, value='Grand Total').font = bf
    ws_p.cell(row=va_gt, column=18).border = tb
    for ci in [19, 20]:
        cl = get_column_letter(ci)
        ws_p.cell(row=va_gt, column=ci, value=f'=SUM({cl}{va_start}:{cl}{va_end})').border = tb
        ws_p.cell(row=va_gt, column=ci).font = bf
        ws_p.cell(row=va_gt, column=ci).number_format = nk
    ws_p.cell(row=va_gt, column=22).border = tb
    for ci in [23, 24, 25]:
        cl = get_column_letter(ci)
        ws_p.cell(row=va_gt, column=ci, value=f'=SUM({cl}{va_start}:{cl}{va_end})').border = tb
        ws_p.cell(row=va_gt, column=ci).font = bf
        ws_p.cell(row=va_gt, column=ci).number_format = nk

    # ========== SECTION 4: Department (A-C below source) ==========
    dept_start = gt1 + 3
    ws_p.cell(row=dept_start, column=1, value='Department').font = hf
    ws_p.cell(row=dept_start, column=1).fill = hfill; ws_p.cell(row=dept_start, column=1).border = tb
    ws_p.cell(row=dept_start, column=2, value='Day').font = hf
    ws_p.cell(row=dept_start, column=2).fill = hfill; ws_p.cell(row=dept_start, column=2).border = tb
    c = ws_p.cell(row=dept_start, column=3, value='Night')
    c.font = Font(bold=True); c.fill = yfill; c.border = tb

    for di, (dept, day, night) in enumerate(DEPT_DATA):
        r = dept_start + 1 + di
        ws_p.cell(row=r, column=1, value=dept).border = tb
        ws_p.cell(row=r, column=2).border = tb
        c = ws_p.cell(row=r, column=3)
        c.border = tb
        c.fill = yfill

    total_r = dept_start + 1 + len(DEPT_DATA)
    ws_p.cell(row=total_r, column=1, value='Total').font = bf
    ws_p.cell(row=total_r, column=1).border = tb
    ws_p.cell(row=total_r, column=2, value=f'=SUM(B{dept_start+1}:B{total_r-1})').font = bf
    ws_p.cell(row=total_r, column=2).border = tb
    ws_p.cell(row=total_r, column=3, value=f'=SUM(C{dept_start+1}:C{total_r-1})').font = bf
    ws_p.cell(row=total_r, column=3).border = tb

    for ci in range(1, 26):
        ws_p.column_dimensions[get_column_letter(ci)].width = 22

    # ========== Tab 4: Formulas & Logic ==========
    ws_f = wb.create_sheet('Formulas & Logic')
    title_font = Font(bold=True, size=14, color='2F5496')
    sec_font = Font(bold=True, size=12, color='2F5496')
    hdr_font = Font(bold=True, size=11)
    ws_f.column_dimensions['A'].width = 8
    ws_f.column_dimensions['B'].width = 35
    ws_f.column_dimensions['C'].width = 55
    ws_f.column_dimensions['D'].width = 55

    r = 1
    ws_f.cell(row=r, column=2, value='IDR Projection Tool - Formulas & Logic').font = title_font
    r += 2

    ws_f.cell(row=r, column=2, value='Parameters Used').font = sec_font
    r += 1
    ws_f.cell(row=r, column=2, value='CF Multiplier (D2)').font = hdr_font
    ws_f.cell(row=r, column=3, value=f'{cf_value} - Produces {cf_value}x of projected quantity as buffer')
    r += 1
    ws_f.cell(row=r, column=2, value='Working Days (F2)').font = hdr_font
    ws_f.cell(row=r, column=3, value=f'{working_days} - Number of factory operating days in the month')
    r += 2

    ws_f.cell(row=r, column=2, value='Tab 1: Projection').font = sec_font
    r += 1
    ws_f.cell(row=r, column=2, value='Content')
    ws_f.cell(row=r, column=3, value='Raw input data copied as-is from the uploaded Projection file')
    r += 1
    ws_f.cell(row=r, column=2, value='Purpose')
    ws_f.cell(row=r, column=3, value='Preserves original data for reference and audit trail')
    r += 2

    ws_f.cell(row=r, column=2, value='Tab 2: IDR Projection').font = sec_font
    r += 1
    ws_f.cell(row=r, column=2, value='Filter Applied')
    ws_f.cell(row=r, column=3, value='Only rows where Origin = "Indore" are included')
    r += 1
    ws_f.cell(row=r, column=2, value='Columns G-H Added')
    ws_f.cell(row=r, column=3, value='Packaging Type & Method derived via VLOOKUP on Item Name from IDR Plan Lookup file')
    r += 1
    ws_f.cell(row=r, column=2, value='Column Order')
    ws_f.cell(row=r, column=3, value='Item Group, Item Name, Item Parent, Conversion Factor, Projection Units, Total KGs, Packaging Type, Packaging Method, Item Type, ...')
    r += 2

    ws_f.cell(row=r, column=2, value='Tab 3: Pivot - Section 1 (Source Summary, Col A-H)').font = sec_font
    r += 1
    formulas_s1 = [
        ('Column A: Source', 'Grouped by Packaging Method from IDR Projection'),
        ('Column B: Projected Units', 'SUM of Projection Units per Packaging Method'),
        ('Column C: Projected Kg', 'SUM of Total KGs per Packaging Method'),
        ('Column D: 1.3 CF', '= B * $D$2 (Projected Units x CF Multiplier)'),
        ('Column E: Sum of Pro (1.3 CF) Kg', '= C * $D$2 (Projected Kg x CF Multiplier)'),
        ('Column F: Per Day PCS Prod.', '= D / $F$2 (CF Units / Working Days)'),
        ('Column G: No. of PCS Prod. (per shift)', 'Shift capacity per source (hardcoded reference values)'),
        ('Column H: No of Cycles', '= F / G (Per Day Production / Shift Capacity)'),
    ]
    for label, desc in formulas_s1:
        ws_f.cell(row=r, column=2, value=label).font = hdr_font
        ws_f.cell(row=r, column=3, value=desc)
        r += 1

    r += 1
    ws_f.cell(row=r, column=2, value='Shift Capacity Reference Values (Col G)').font = sec_font
    r += 1
    ws_f.cell(row=r, column=2, value='Source').font = hdr_font
    ws_f.cell(row=r, column=3, value='Capacity Formula').font = hdr_font
    r += 1
    for source, cap in SHIFT_CAPACITY.items():
        ws_f.cell(row=r, column=2, value=source)
        ws_f.cell(row=r, column=3, value=str(cap))
        r += 1

    r += 1
    ws_f.cell(row=r, column=2, value='Tab 3: Pivot - Section 2a (FW Items, Col K-O)').font = sec_font
    r += 1
    formulas_fw = [
        ('Column K: Item Name', 'Filtered where Packaging Method = "FW", grouped by Item Name'),
        ('Column L: Projected Units', 'SUM of Projection Units per Item Name'),
        ('Column M: Projected Kg', 'SUM of Total KGs per Item Name'),
        ('Column N: 1.3 CF', '= M * $N$2 (Projected Kg x CF Multiplier)'),
        ('Column O: Per Day Production Kg', '= N / $O$2 (CF Kg / Working Days)'),
    ]
    for label, desc in formulas_fw:
        ws_f.cell(row=r, column=2, value=label).font = hdr_font
        ws_f.cell(row=r, column=3, value=desc)
        r += 1

    r += 1
    ws_f.cell(row=r, column=2, value='Tab 3: Pivot - Section 2b (Multiple Items = Pace-FFS + Perfect-FFS, Col K-O)').font = sec_font
    r += 1
    ws_f.cell(row=r, column=2, value='Filter')
    ws_f.cell(row=r, column=3, value='Packaging Method IN ("Pace - FFS", "Perfect - FFS") combined as "(Multiple Items)"')
    r += 1
    ws_f.cell(row=r, column=2, value='Same formulas as FW section with separate CF/Days parameter references')
    r += 2

    ws_f.cell(row=r, column=2, value='Tab 3: Pivot - Section 3 (Value Added, Col R-Y)').font = sec_font
    r += 1
    formulas_va = [
        ('Column R: Item Name', 'Filtered where Item Type = "Value Added", grouped by Item Parent'),
        ('Column S: Projected Kg', 'SUM of Total KGs per Item Parent'),
        ('Column T: 1.3 CF', '= S * $T$2 (Projected Kg x CF Multiplier)'),
        ('Column U: Per Day Kg Prod', '= T / $U$2 (CF Kg / Working Days)'),
        ('Column V: Roasting Item (%)', 'User-fillable: % of item that requires roasting'),
        ('Column W: Roasting Kg', '= U * V (Per Day Kg x Roasting %)'),
        ('Column X: Oven Cycle (SKU Level)', 'User-fillable: Oven cycle capacity in Kg per cycle'),
        ('Column Y: No of Oven Cycle', '= IF(X>0, W/X, 0) (Roasting Kg / Oven Cycle Capacity)'),
    ]
    for label, desc in formulas_va:
        ws_f.cell(row=r, column=2, value=label).font = hdr_font
        ws_f.cell(row=r, column=3, value=desc)
        r += 1

    r += 1
    ws_f.cell(row=r, column=2, value='Tab 3: Pivot - Section 4 (Department Manpower, Col A-C)').font = sec_font
    r += 1
    ws_f.cell(row=r, column=2, value='Day & Night columns')
    ws_f.cell(row=r, column=3, value='Empty cells for manual entry of department-wise shift headcount')
    r += 1
    ws_f.cell(row=r, column=2, value='Total Row')
    ws_f.cell(row=r, column=3, value='= SUM of Day column, SUM of Night column (auto-calculated)')
    r += 1
    ws_f.cell(row=r, column=2, value='Night column highlight')
    ws_f.cell(row=r, column=3, value='Yellow background indicates Night shift cells')
    r += 2

    ws_f.cell(row=r, column=2, value='Processing Pipeline Summary').font = sec_font
    r += 1
    steps = [
        ('Step 1', 'Read Projection file (auto-detects sheet: "Projection", "Query Report", or first sheet)'),
        ('Step 2', 'Read IDR Plan Lookup file for Item Name -> Packaging Type/Method mapping'),
        ('Step 3', 'Build IDR Projection: reorder columns, add Packaging Type & Method via lookup'),
        ('Step 4', 'Filter: keep only rows where Origin = "Indore"'),
        ('Step 5', 'Generate Pivot Source Summary: group by Packaging Method, apply CF & Working Days formulas'),
        ('Step 6', 'Generate FW & FFS item breakdowns with per-day production calculations'),
        ('Step 7', 'Generate Value Added breakdown with roasting formula placeholders'),
        ('Step 8', 'Department manpower table with empty Day/Night cells for manual entry'),
    ]
    for label, desc in steps:
        ws_f.cell(row=r, column=2, value=label).font = hdr_font
        ws_f.cell(row=r, column=3, value=desc)
        r += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/process', methods=['POST'])
def process():
    if 'this_month' not in request.files or 'lookup_file' not in request.files:
        return jsonify({'error': 'Both files are required'}), 400

    this_month = request.files['this_month']
    lookup_f = request.files['lookup_file']
    cf_value = float(request.form.get('cf_value', 1.3))
    working_days = int(request.form.get('working_days', 25))

    this_buf = io.BytesIO(this_month.read())
    lookup_buf = io.BytesIO(lookup_f.read())

    try:
        df_proj, df_idr, pivot_source, pivot_fw, pivot_ffs, pivot_va = process_files(this_buf, lookup_buf, cf_value, working_days)
    except Exception as e:
        return jsonify({'error': f'Processing error: {str(e)}'}), 500

    output = build_output_excel(df_proj, df_idr, pivot_source, pivot_fw, pivot_ffs, pivot_va, cf_value, working_days)

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    tmp.write(output.read())
    tmp.close()
    app.config['LAST_OUTPUT'] = tmp.name

    # Build VA data with roasting for dashboard
    va_dashboard = []
    for _, row in pivot_va.iterrows():
        item = row['Item Name']
        kg = row['Projected Kg']
        cf_kg = kg * cf_value
        per_day = cf_kg / working_days
        va_dashboard.append({
            'Item Name': item, 'Projected Kg': round(kg, 2),
            '1.3 CF': round(cf_kg, 2), 'Per Day Kg Prod': round(per_day, 0),
            'Roasting Item (%)': '', 'Roasting Kg': '',
            'Oven Cycle': '', 'No of Oven Cycle': ''
        })

    # Build FFS dashboard data
    ffs_dashboard = []
    for _, row in pivot_ffs.iterrows():
        cf_val = row['Projected Kg'] * cf_value
        ffs_dashboard.append({
            'Item Name': row['Item Name'],
            'Projected Units': int(row['Projected Units']),
            'Projected Kg': round(row['Projected Kg'], 2),
            '1.3 CF': round(cf_val, 2),
            'Per Day Production Kg': round(cf_val / working_days, 0)
        })

    summary = {
        'total_items': len(df_idr),
        'total_raw_items': len(df_proj),
        'total_projection_units': int(df_idr['Projection Units'].sum()),
        'total_kgs': round(float(df_idr['Total KGs'].sum()), 2),
        'unique_items': int(df_idr['Item Name'].nunique()),
        'unique_customers': int(df_idr['Customer'].nunique()),
        'cf_value': cf_value,
        'working_days': working_days,
        'item_groups': df_idr['Item Group'].value_counts().head(15).to_dict(),
        'customer_groups': df_idr['Customer Group'].value_counts().head(10).to_dict(),
        'packaging_methods': df_idr['Packaging Method'].value_counts().head(10).to_dict(),
        'packaging_types': df_idr['Packaging Type'].value_counts().head(10).to_dict(),
        'item_types': df_idr['Item Type'].value_counts().to_dict(),
        'origins': df_idr['Origin'].value_counts().head(10).to_dict(),
        'top_items_by_kg': df_idr.groupby('Item Name')['Total KGs'].sum().nlargest(10).to_dict(),
        'kam_summary': df_idr.groupby('Key Account Manager')['Total KGs'].sum().nlargest(10).to_dict(),
        'pivot_source': pivot_source.to_dict(orient='records'),
        'pivot_fw': pivot_fw.to_dict(orient='records'),
        'pivot_ffs': ffs_dashboard,
        'pivot_va': va_dashboard,
        'idr_preview': json.loads(df_idr.head(50).to_json(orient='records', default_handler=str)),
        'pkg_match_rate': round(
            (df_idr['Packaging Type'].notna() & (df_idr['Packaging Type'] != '') & (df_idr['Packaging Type'] != 'nan')).sum() / len(df_idr) * 100, 1
        ) if len(df_idr) > 0 else 0,
        'dept_data': [{'Department': d, 'Day': '', 'Night': ''} for d, dy, n in DEPT_DATA],
    }

    return jsonify(summary)


@app.route('/download')
def download():
    output_path = app.config.get('LAST_OUTPUT')
    if output_path and os.path.exists(output_path):
        return send_file(output_path, as_attachment=True, download_name='IDR_Projection_Output.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    return jsonify({'error': 'No output file found. Process files first.'}), 404


if __name__ == '__main__':
    app.run(debug=True, port=5000)
